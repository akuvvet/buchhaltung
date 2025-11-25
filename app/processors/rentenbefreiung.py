from __future__ import annotations

from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Optional, Tuple

from io import BytesIO

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
import subprocess
import shutil


DATE_FMT = "%d.%m.%Y"


def parse_date_ddmmyyyy(value: str, *, allow_today_default: bool = False) -> str:
	value = (value or "").strip()
	if not value and allow_today_default:
		return datetime.today().strftime(DATE_FMT)
	try:
		dt = datetime.strptime(value, DATE_FMT)
		return dt.strftime(DATE_FMT)
	except ValueError:
		raise ValueError("Bitte Datum im Format TT.MM.JJJJ angeben (z. B. 25.11.2025).")


def _fill_excel_and_export_pdf(
	excel_path: Path,
	familienname: str,
	vorname: str,
	rvnr: str,
	ort: str,
	aktuelles_datum: str,
	beginn_befreiung: str,
	signature_path: Optional[Path] = None,
) -> bytes:
	# Workbook laden
	wb = load_workbook(filename=str(excel_path))
	ws = wb.worksheets[0]

	# Einträge in vorgegebene Zellen
	ws["C10"].value = familienname
	ws["C10"].font = Font(bold=True)
	ws["C12"].value = vorname
	ws["C12"].font = Font(bold=True)
	ws["C14"].value = rvnr
	ws["C14"].font = Font(bold=True)

	ort_komma_datum = f"{ort}, {aktuelles_datum}"
	ws["B26"].value = ort_komma_datum
	ws["B26"].font = Font(bold=True)
	ws["B40"].value = ort_komma_datum
	ws["B40"].font = Font(bold=True)

	ws["E36"].value = aktuelles_datum
	ws["E36"].font = Font(bold=True)
	ws["D38"].value = beginn_befreiung
	ws["D38"].font = Font(bold=True)

	# Unterschriftsbild optional bei Zelle D27 verankern
	if signature_path is not None and signature_path.exists():
		try:
			img = XLImage(str(signature_path))
			# Maximal 5 cm Breite und 2 cm Höhe, proportional
			# Umrechnung: 1 inch = 2.54 cm, 96 px/inch
			DPI = 96.0
			max_w_px = int((5.0 / 2.54) * DPI)
			max_h_px = int((2.0 / 2.54) * DPI)
			cur_w = int(getattr(img, "width", max_w_px) or max_w_px)
			cur_h = int(getattr(img, "height", max_h_px) or max_h_px)
			if cur_w > 0 and cur_h > 0:
				scale = min(max_w_px / cur_w, max_h_px / cur_h, 1.0)
				if scale < 1.0:
					img.width = int(cur_w * scale)
					img.height = int(cur_h * scale)
			# Verankerung an D27
			img.anchor = "D27"
			ws.add_image(img)
		except Exception:
			# Wenn Bild fehlschlägt, weiter ohne Bild
			pass

	# In Bytes speichern (xlsx)
	bio = BytesIO()
	wb.save(bio)
	bio.seek(0)
	return bio.getvalue()


def export_rentenbefreiung_xlsx(
	excel_bytes: bytes,
	familienname: str,
	vorname: str,
	rvnr: str,
	ort: str,
	aktuelles_datum: str,
	beginn_befreiung: str,
	signature_bytes: Optional[bytes] = None,
) -> Tuple[bytes, str]:
	"""
	Nimmt das Excel-Template als Bytes entgegen, befüllt es per openpyxl und liefert XLSX-Bytes zurück.
	"""
	familienname_clean = (familienname or "").strip()
	vorname_clean = (vorname or "").strip()
	if not familienname_clean:
		raise ValueError("Familienname darf nicht leer sein.")
	if not vorname_clean:
		raise ValueError("Vorname darf nicht leer sein.")
	if not (rvnr or "").strip():
		raise ValueError("Rentenversicherungsnr darf nicht leer sein.")
	ort_clean = (ort or "").strip() or "Solingen"
	aktuelles_datum_norm = parse_date_ddmmyyyy(aktuelles_datum, allow_today_default=True)
	beginn_befreiung_norm = parse_date_ddmmyyyy(beginn_befreiung, allow_today_default=False)

	with TemporaryDirectory(prefix="rentenbefreiung_") as tmpdir:
		tmp_dir = Path(tmpdir)
		excel_path = tmp_dir / "vorlage.xlsx"
		excel_path.write_bytes(excel_bytes)
		signature_path: Optional[Path] = None
		if signature_bytes:
			signature_path = tmp_dir / "signature.png"
			signature_path.write_bytes(signature_bytes)

		result_xlsx_bytes = _fill_excel_and_export_pdf(
			excel_path=excel_path,
			familienname=familienname_clean,
			vorname=vorname_clean,
			rvnr=(rvnr or "").strip(),
			ort=ort_clean,
			aktuelles_datum=aktuelles_datum_norm,
			beginn_befreiung=beginn_befreiung_norm,
			signature_path=signature_path,
		)

		filename = f"Rentenbefreiung_{familienname_clean}_{vorname_clean}.xlsx"
		return result_xlsx_bytes, filename


def export_rentenbefreiung_pdf(
	excel_bytes: bytes,
	familienname: str,
	vorname: str,
	rvnr: str,
	ort: str,
	aktuelles_datum: str,
	beginn_befreiung: str,
	signature_bytes: Optional[bytes] = None,
) -> Tuple[bytes, str]:
	"""
	Erzeugt zunächst eine ausgefüllte XLSX-Datei (openpyxl) und konvertiert diese anschließend
	mittels LibreOffice (soffice --headless) nach PDF. Erfordert, dass LibreOffice/soffice
	auf dem Server installiert ist.
	"""
	# Zuerst XLSX erzeugen
	xlsx_bytes, _xlsx_name = export_rentenbefreiung_xlsx(
		excel_bytes=excel_bytes,
		familienname=familienname,
		vorname=vorname,
		rvnr=rvnr,
		ort=ort,
		aktuelles_datum=aktuelles_datum,
		beginn_befreiung=beginn_befreiung,
		signature_bytes=signature_bytes,
	)

	# Pfad zu soffice finden
	soffice = shutil.which("soffice") or shutil.which("libreoffice") or shutil.which("lowriter")
	if not soffice:
		raise RuntimeError("LibreOffice (soffice) nicht gefunden. Bitte installieren, um PDF zu erzeugen.")

	with TemporaryDirectory(prefix="rentenbefreiung_pdf_") as tmpdir:
		tmp_dir = Path(tmpdir)
		in_xlsx = tmp_dir / "rentenbefreiung.xlsx"
		out_dir = tmp_dir
		in_xlsx.write_bytes(xlsx_bytes)

		# Nach PDF konvertieren
		cmd = [
			soffice,
			"--headless",
			"--convert-to",
			"pdf",
			"--outdir",
			str(out_dir),
			str(in_xlsx),
		]
		res = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
		if res.returncode != 0:
			raise RuntimeError(f"LibreOffice-Konvertierung fehlgeschlagen: {res.stderr or res.stdout}")

		out_pdf = out_dir / "rentenbefreiung.pdf"
		if not out_pdf.exists():
			# LibreOffice nutzt normalerweise denselben Basenamen
			# Fallback: irgendeine erzeugte PDF nehmen
			candidates = list(out_dir.glob("*.pdf"))
			if not candidates:
				raise RuntimeError("PDF nicht erzeugt.")
			out_pdf = candidates[0]

		pdf_bytes = out_pdf.read_bytes()

		filename = f"Rentenbefreiung_{(familienname or '').strip()}_{(vorname or '').strip()}.pdf"
		return pdf_bytes, filename


