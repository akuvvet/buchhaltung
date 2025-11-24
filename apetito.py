import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
import pyperclip
import glob
import time

# ... (all your other functions: process_tours_table, process_excel, copy_to_clipboard, suggest_file_with_date) ...
# These functions are assumed to be the same as in your last provided code. I will paste them for completeness.

# Hinzufügen der neuen Funktion zur Bearbeitung der Touren-Tabelle
def process_tours_table(workbook):
    if "touren" not in workbook.sheetnames:
        workbook.create_sheet("touren")
    tours_sheet = workbook["touren"]
    
    # Ersetzen Sie dies durch den tatsächlichen Namen Ihrer "ag-grid" Tabelle, falls anders
    if "ag-grid" not in workbook.sheetnames:
        print("Warning: Sheet 'ag-grid' not found. Skipping tours table processing.")
        return # Exit if dependent sheet is missing
    ag_grid_sheet = workbook["ag-grid"] 

    # Ermitteln des aktuellen Datums
    current_date = datetime.datetime.now().strftime("%Y-%m-%d")

    # Benenne Spaltenüberschriften
    tours_sheet.cell(1, 1, "Tour")
    tours_sheet.cell(1, 2, f"TG ({current_date})")

    # Kopiere Spalte A von "ag-grid" nach "touren" und sammle die Tour-Werte
    tour_values = []
    for row in range(2, ag_grid_sheet.max_row + 1):  # Beginne bei Zeile 2, um Überschriften zu ignorieren
        tour_value = ag_grid_sheet.cell(row, 1).value
        if tour_value:
            tour_values.append(tour_value)

    # Entfernen von Duplikaten und Sortieren
    unique_tours = sorted(list(set(tour_values)))
    
    if tours_sheet.max_row >= 2:
        tours_sheet.delete_rows(2, tours_sheet.max_row -1) 

    for i, tour in enumerate(unique_tours, 2):  # Beginne bei Zeile 2 für Überschriften
        tours_sheet.cell(i, 1, tour)

    # Zählen der Kundennummern pro Tour
    tour_customer_count = {}
    for row in range(2, ag_grid_sheet.max_row + 1):  # Beginne bei Zeile 2
        tour = ag_grid_sheet.cell(row, 1).value  # Spalte A für Touren
        customer = ag_grid_sheet.cell(row, 8).value  # Spalte H für Kundennummern
        if tour and customer: # Ensure customer is not None or empty before counting
            tour_customer_count[tour] = tour_customer_count.get(tour, 0) + 1
            
    for row_idx in range(2, len(unique_tours) + 2): 
        tour = tours_sheet.cell(row_idx, 1).value
        if tour: # Ensure tour value exists before trying to get count
            tours_sheet.cell(row_idx, 2, tour_customer_count.get(tour, 0))


def process_excel(file_path, save_directory):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    process_tours_table(workbook)

    blue_fill = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    light_gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    col_ac_idx = 29
    col_ad_idx = 30
    col_ae_idx = 31
    col_af_idx = 32
    col_ag_idx = 33
    col_ah_idx = 34

    for row in range(1, sheet.max_row + 1):
        sheet.cell(row, col_ac_idx).fill = blue_fill
    sheet.cell(1, col_ad_idx).fill = blue_fill

    sheet.cell(1, col_ae_idx).fill = green_fill
    sheet.cell(1, col_af_idx).fill = green_fill

    sheet.cell(1, col_ag_idx).fill = light_gray_fill
    sheet.cell(1, col_ah_idx).fill = light_gray_fill

    values_to_filter = ["D009", "D090", "D091", "D092", "D093", "D094", "D096", "D095", "D208", "D251", "D270", "D271", "D291", "D292", "SCD12", "SCD13"]
    sheet.auto_filter.add_filter_column(0, values_to_filter)
    
    new_filter_last_col_letter = get_column_letter(28) 
    filter_range = f"A1:{new_filter_last_col_letter}{sheet.max_row}"
    sheet.auto_filter.ref = filter_range
    
    subtotal_source_col_letter = get_column_letter(col_ac_idx)
    current_max_row = sheet.max_row
    if current_max_row < 2 : current_max_row = 2 
    subtotal_range_end_row = max(current_max_row, 2000) 
    sheet.cell(1, col_ad_idx).value = f'=SUBTOTAL(9,{subtotal_source_col_letter}2:{subtotal_source_col_letter}{subtotal_range_end_row})'

    sheet.cell(1, col_ae_idx).value = 'Adressen'

    sheet.cell(1, col_af_idx).value = '=SUMPRODUCT(--(FREQUENCY(COLUMN(1:1175),SUBTOTAL(3,INDIRECT("H"&ROW(2:1175)))*MATCH(H2:H1175&"",H2:H1175&"",0))>0))-1'
    
    sheet.cell(1, col_ag_idx).value = 'Touren'

    sheet.cell(1, col_ah_idx).value = '=SUMPRODUCT(--(FREQUENCY(COLUMN(1:1),SUBTOTAL(3,INDIRECT("A"&ROW(2:1175)))*MATCH(A2:A1175&"",A2:A1175&"",0))>0))-1'

    sheet.column_dimensions['A'].width = 6.5
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 6
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 8
    sheet.column_dimensions['G'].width = 15
    sheet.column_dimensions['H'].width = 15

    col_i_idx = 9
    col_j_idx = 10
    sheet.cell(row=1, column=col_i_idx).value = "Ablage"
    sheet.column_dimensions[get_column_letter(col_i_idx)].width = 8
    sheet.cell(row=1, column=col_j_idx).value = "Schlüssel"
    sheet.column_dimensions[get_column_letter(col_j_idx)].width = 8

    col_k_idx = 11 
    col_l_idx = 12 
    sheet.column_dimensions[get_column_letter(col_k_idx)].width = 15
    sheet.column_dimensions[get_column_letter(col_l_idx)].width = 10
    
    comment_target_col_idx = col_l_idx 
    comment_source_start_col_idx = 13  
    comment_source_end_col_idx = 22    

    for row_num in range(2, sheet.max_row + 1):
        comment_text = ""
        for col_idx_comment_source in range(comment_source_start_col_idx, comment_source_end_col_idx + 1):
            cell_value = sheet.cell(row_num, col_idx_comment_source).value
            if cell_value is not None:
                column_letter_for_comment = get_column_letter(col_idx_comment_source)
                comment_text += f"{column_letter_for_comment}{row_num}: {cell_value}\n"

        if comment_text:
            comment = Comment(comment_text.strip(), "System")
            comment.width = 200
            comment.height = 400
            sheet.cell(row_num, comment_target_col_idx).comment = comment

    for cell in sheet["G"]:
        cell.number_format = '0'

    for col_idx_hide in range(13, 29): 
        sheet.column_dimensions[get_column_letter(col_idx_hide)].hidden = True

    sheet.cell(1, col_ac_idx).value = "Menü" 

    starts_with_list = ["1 4", "2 4", "3 4", "4 4", "5 4", "6 4"]
    source_count_col_idx = col_l_idx 
    target_count_col_idx = col_ac_idx 

    for row_num in range(2, sheet.max_row + 1):
        cell_value_j = sheet.cell(row_num, source_count_col_idx).value
        count = 0 
        if cell_value_j and isinstance(cell_value_j, str):
            lines = cell_value_j.splitlines()
            for line in lines:
                for start_pattern in starts_with_list:
                    if line.strip().startswith(start_pattern):
                        try:
                            count += int(start_pattern.split()[0])
                        except (ValueError, IndexError):
                            pass 
        
        if count > 0:
            sheet.cell(row_num, target_count_col_idx).value = count
        else: 
            sheet.cell(row_num, target_count_col_idx).value = None

    current_date_str = datetime.datetime.now().strftime("%Y%m%d")
    modified_file_name = f"{current_date_str}.xlsx"
    modified_file_path = os.path.join(save_directory, modified_file_name)
    
    if not os.path.exists(save_directory):
        os.makedirs(save_directory)
        
    workbook.save(modified_file_path)
    return modified_file_path


def copy_to_clipboard(sheet):
    clipboard_data = ""
    for row_num in range(2, sheet.max_row + 1):
        if sheet.cell(row_num, 6).value in [1, 2]:  
            row_data = []
            for col_idx in range(1, 12):  # Changed to 12 to include column K (11 columns)
                cell_val = sheet.cell(row_num, col_idx).value
                if col_idx == 11:  # Column K
                    if cell_val and "LHK" in str(cell_val):
                        row_data.append("LHK")
                    else:
                        row_data.append("MS")
                else:
                    row_data.append(str(cell_val) if cell_val is not None else "")
            clipboard_data += "\t".join(row_data) + "\n"
    
    if clipboard_data:
        pyperclip.copy(clipboard_data)
        print("Data copied to clipboard.")
    else:
        print("No data matching criteria to copy to clipboard.")


def suggest_file_with_date(directory, date_str):
    pattern = os.path.join(directory, f"*{date_str}*.xlsx")
    files = glob.glob(pattern)
    if files:
        files.sort(key=os.path.getmtime, reverse=True) 
        return files[0]
    return None


def show_auto_close_message(title, message, timeout=5000):
    root = tk.Tk()
    root.withdraw()
    
    # Create a new window
    popup = tk.Toplevel(root)
    popup.title(title)
    
    # Center the window
    popup.geometry("+%d+%d" % (root.winfo_x() + 50, root.winfo_y() + 50))
    
    # Add message
    label = tk.Label(popup, text=message, padx=20, pady=20)
    label.pack()
    
    # Function to close the window and exit script
    def close_popup():
        popup.destroy()
        root.destroy()
        # Schedule script exit after 2 seconds
        time.sleep(2)
        os._exit(0)
    
    # Schedule the window to close after timeout
    popup.after(timeout, close_popup)
    
    # Make the window appear on top
    popup.lift()
    popup.attributes('-topmost', True)
    
    # Start the main loop
    root.mainloop()


def main():
    current_date_str = datetime.datetime.now().strftime("%Y%m%d")
    save_directory = 'O:\\apetito\\apetitotelematik' 

    suggested_file = suggest_file_with_date(save_directory, current_date_str)

    root = tk.Tk()
    root.withdraw()

    initial_dir_val = save_directory
    initial_file_val = ""
    if suggested_file:
        # Ensure suggested_file is a valid path and os.path functions can handle it
        try:
            candidate_dir = os.path.dirname(suggested_file)
            candidate_file = os.path.basename(suggested_file)
            
            # Check if the derived path components actually form an existing file path
            # This is a defense against glob returning something unusual or a race condition.
            if os.path.exists(os.path.join(candidate_dir, candidate_file)):
                initial_dir_val = candidate_dir
                initial_file_val = candidate_file
            else:
                print(f"Warning: Suggested file '{suggested_file}' (basename: '{candidate_file}') "
                      f"in directory '{candidate_dir}' was not confirmed by os.path.exists. "
                      f"Falling back to default directory without pre-selected file.")
                # Keep initial_dir_val as save_directory and initial_file_val as ""
        except Exception as e:
            print(f"Error processing suggested_file '{suggested_file}': {e}. "
                  f"Falling back to default directory.")
            # Keep initial_dir_val as save_directory and initial_file_val as ""

    file_path = None # Initialize file_path
    print(f"Attempting to open file dialog with: initialdir='{initial_dir_val}', initialfile='{initial_file_val}'")
    try:
        file_path = filedialog.askopenfilename(
            title="Excel-Datei auswählen",
            initialdir=initial_dir_val,
            initialfile=initial_file_val,
            filetypes=[("Excel files", "*.xlsx")]
        )
    except tk.TclError as e:
        print(f"Tkinter TclError occurred with pre-selected file: {e}")
        print("Retrying file dialog without pre-selected file and using default save directory.")
        try:
            file_path = filedialog.askopenfilename(
                title="Excel-Datei auswählen (Retry)",
                initialdir=save_directory, # Fallback to the base save_directory
                filetypes=[("Excel files", "*.xlsx")]
            )
        except Exception as e_retry:
            print(f"Retry file dialog also failed: {e_retry}")
            # Fall through, file_path will be None or the result of the first (failed) dialog
    except Exception as e_other:
        print(f"An unexpected error occurred during file dialog: {e_other}")
        # Fall through, file_path might be None

    if file_path:
        try:
            print(f"Processing file: {file_path}")
            modified_file_path = process_excel(file_path, save_directory)
            print(f"Modified file saved at: {modified_file_path}")

            workbook_for_clipboard = openpyxl.load_workbook(modified_file_path)
            active_sheet_for_clipboard = workbook_for_clipboard.active 
            copy_to_clipboard(active_sheet_for_clipboard)

            os.startfile(modified_file_path)
            time.sleep(3)  # Wait for 5 seconds
            show_auto_close_message("Information", "Neukunden wurden in die Zwischenablage kopiert")
        except Exception as e_process:
            print(f"An error occurred during Excel processing or file operations: {e_process}")
            # import tkinter.messagebox # Keep this if you want a GUI error too
            # tkinter.messagebox.showerror("Processing Error", f"An error occurred:\n{e_process}")
    else:
        print("No file selected or file dialog failed.")

if __name__ == "__main__":
    main()